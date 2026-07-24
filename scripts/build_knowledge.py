from pathlib import Path
import json
from openpyxl import load_workbook
from docx import Document


SOURCE = Path(r'C:\Users\Admin\Desktop\Скрипты продаж')
OUT = Path('data/knowledge.json')
CURATED = Path('data/curated_rules.json')
items = []


def add(text, source, section):
    text = str(text).strip()
    if len(text) >= 20:
        items.append({'text': text, 'source': source, 'section': section})


backup = json.loads((SOURCE / 'backup_alina_2026-06-29 2.json').read_text(encoding='utf-8-sig'))
for stage in backup:
    for src in stage.get('sources', []):
        for situation in src.get('situations', []):
            for version in situation.get('versions', []):
                for text in version.get('texts', []):
                    add(text, 'backup_alina_2026-06-29 2.json', f"{stage['name']} / {src['name']} / {situation['name']}")

for filename in ['скрипты продаж.xlsx', 'ответы на вопросы клиентов.xlsx', 'скрипты новые.xlsx']:
    wb = load_workbook(SOURCE / filename, read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str): add(value, filename, ws.title)

doc = Document(SOURCE / 'скрипт звонка.docx')
for p in doc.paragraphs: add(p.text, 'скрипт звонка.docx', 'Основной текст')

if CURATED.exists():
    for item in json.loads(CURATED.read_text(encoding='utf-8-sig')):
        add(item['text'], 'curated_rules.json', item.get('section', 'Проверенные правила'))

dedup = {}
for item in items:
    key = ' '.join(item['text'].lower().split())
    dedup.setdefault(key, item)
OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(list(dedup.values()), ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Knowledge items: {len(dedup)} -> {OUT}')
